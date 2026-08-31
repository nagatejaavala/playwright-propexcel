#!/usr/bin/env python3
"""Update PropExcel_Test_Cases.xlsx to match current Playwright spec files."""
from __future__ import annotations

import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "PropExcel_Test_Cases.xlsx"

HEADERS = [
    "Test ID",
    "Title",
    "Step #",
    "Actor",
    "Action",
    "Expected Result",
    "Priority",
    "Type",
    "Precondition",
]


def replace_in_sheet(ws, mapping: dict[str, str]) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                val = cell.value
                for old, new in mapping.items():
                    val = val.replace(old, new)
                cell.value = val


def write_test_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    title: str,
    test_id: str,
    precondition: str,
    steps: list[tuple[str, str, str]],
    script: str,
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"], ws["B3"] = "Test ID", test_id
    ws["A4"], ws["B4"] = "Priority", "High"
    ws["A5"], ws["B5"] = "Type", "End-to-end / Regression"
    ws["A6"], ws["B6"] = "Precondition", precondition
    ws["A7"], ws["B7"] = "Script", script

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)

    for i, (actor, action, expected) in enumerate(steps, start=1):
        row = 8 + i
        ws.cell(row=row, column=1, value=test_id)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=i)
        ws.cell(row=row, column=4, value=actor)
        ws.cell(row=row, column=5, value=action)
        ws.cell(row=row, column=6, value=expected)
        ws.cell(row=row, column=7, value="High")
        ws.cell(row=row, column=8, value="Regression")
        ws.cell(row=row, column=9, value=precondition if i == 1 else None)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["I"].width = 60


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    global_replace = {
        "test240": "test",
        "org: test240": "org: test (test@yopmail.com)",
        "YOPmail": "IMAP (Gmail app password)",
        "yopmail": "gmail.com (propexceltest+tenantN)",
        "contact*), email @gmail.com (propexceltest+tenantN)": "tenantN), email propexceltest+tenantN@gmail.com",
    }

    for name in wb.sheetnames:
        replace_in_sheet(wb[name], global_replace)

    # --- Summary ---
    ws = wb["Summary"]
    ws["B5"] = "Existing org: test (test@yopmail.com / Test2026$) | New Org: dynamic org from test-data/org.json"
    ws["B7"] = (
        "tests/Cleanup Scenario.spec.ts, tests/CreateOrganization.spec.ts, "
        "tests/Flow1-ExistingOrganization.spec.ts, tests/Flow2-ExistingOrganization.spec.ts, "
        "tests/Flow1-NewOrganization.spec.ts, tests/Flow2-NewOrganization.spec.ts, "
        "tests/Creating Contacts,Leads,Deals.spec.ts, tests/ExistingContact-ToTenantPayment.spec.ts, "
        "tests/Skip lead flow - tenenat Payment.spec.ts, tests/Direct Lead to Tenant Payment.spec.ts"
    )
    ws["B8"] = (
        "New Org chain: TC-00 Cleanup -> TC-03 CreateOrganization -> TC-04 Flow1-NewOrganization -> "
        "TC-06 Creating Contacts,Leads,Deals -> TC-07 ExistingContact-ToTenantPayment -> "
        "TC-08 Skip lead flow -> TC-09 Direct Lead to Tenant Payment -> TC-05 Flow2-NewOrganization (optional). "
        "Existing org chain: TC-01 Flow1-ExistingOrganization -> TC-02 Flow2-ExistingOrganization."
    )
    ws["B9"] = (
        "Automated platform cleanup, organization registration, CRM contacts/leads/deals, tenant onboarding, "
        "rent collection, service request, vendor billing, tenant invoicing, and Razorpay payment journeys "
        "for both existing org (test) and newly created organizations. Tenant credentials via IMAP."
    )

    overview = [
        ("TC-00", "Delete auto organizations (platform admin). Script: Cleanup Scenario.spec.ts"),
        ("TC-01", "Create property, onboard tenant, rent invoice, Razorpay pay (org: test). Script: Flow1-ExistingOrganization.spec.ts"),
        ("TC-02", "Tenant request -> vendor -> task -> approval -> bill -> invoice -> Razorpay (org: test). Script: Flow2-ExistingOrganization.spec.ts"),
        ("TC-03", "Create new organization via Stripe; save org.json. Script: CreateOrganization.spec.ts"),
        ("TC-04", "New-org Flow 1: Razorpay, GST, Deal Approve, property, tenant, invoice, Razorpay. Script: Flow1-NewOrganization.spec.ts"),
        ("TC-05", "New-org Flow 2: request, vendor, task, bill, invoice, Razorpay. Script: Flow2-NewOrganization.spec.ts"),
        ("TC-06", "Create 4 contacts + 4 leads (tenantN sequential); save crm-contacts-leads.json. Script: Creating Contacts,Leads,Deals.spec.ts"),
        ("TC-07", "Existing CRM contact -> lead/deal -> tenant -> rent invoice paid. Script: ExistingContact-ToTenantPayment.spec.ts"),
        ("TC-08", "Skip lead: existing contact -> Create Deal -> tenant -> rent payment. Script: Skip lead flow - tenenat Payment.spec.ts"),
        ("TC-09", "Direct lead: auto contact -> deal -> tenant -> rent payment. Script: Direct Lead to Tenant Payment.spec.ts"),
    ]
    start_row = 13
    for i, (tc_id, desc) in enumerate(overview):
        ws.cell(row=start_row + i, column=1, value=tc_id)
        ws.cell(row=start_row + i, column=2, value=desc)

    # Clear old extra summary rows if any
    for r in range(start_row + len(overview), 30):
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=None)

    # --- TC-01 targeted fixes ---
    ws1 = wb["TC-01_Flow1"]
    ws1["B6"] = (
        "Existing org Super Admin credentials valid: org test, email test@yopmail.com, password Test2026$. "
        "TEST environment available. GMAIL_APP_PASSWORD set for IMAP."
    )
    for row in ws1.iter_rows(min_row=9, max_row=ws1.max_row):
        action = row[4].value
        if isinstance(action, str):
            if "Login with org test" in action:
                row[4].value = "Login with org test, email test@yopmail.com, password Test2026$"
            if "create a new contact" in action.lower():
                row[4].value = (
                    "Go to CRM Contacts and create contact (or reuse if duplicate) with tenantN name, "
                    "propexceltest+tenantN@gmail.com, India mobile, Indian nationality"
                )
            if "create lead" in action.lower() and row[3].value == "Super Admin":
                row[4].value = (
                    "Open contact and create lead (or open existing lead if duplicate email/phone exists)"
                )
            if row[2].value == 1:
                row[8].value = ws1["B6"].value

    # --- TC-02 targeted fixes ---
    ws2 = wb["TC-02_Flow2"]
    ws2["B6"] = "TC-01 completed; test-data/tenant.json exists with orgId test and valid tenant credentials."
    for row in ws2.iter_rows(min_row=9, max_row=ws2.max_row):
        action = row[4].value
        if isinstance(action, str):
            if "Login using admin credentials" in action:
                row[4].value = "Login as Super Admin: org test, email test@yopmail.com, password Test2026$"
            if "assign the created vendor" in action:
                row[4].value = (
                    "Add task with status Open, random priority, assign created vendor "
                    "(fallback: Super Admin if vendor not in dropdown yet)"
                )
            if row[2].value == 1:
                row[8].value = ws2["B6"].value

    # --- TC-06 email fix ---
    ws6 = wb["TC-06_ContactsLeadsDeals"]
    for row in ws6.iter_rows(min_row=9, max_row=ws6.max_row):
        if isinstance(row[4].value, str) and "create 4 contacts" in row[4].value:
            row[4].value = (
                "Go to CRM -> Contacts and create 4 contacts (tenantN sequential): "
                "fullName tenantN, email propexceltest+tenantN@gmail.com, India mobile, Indian nationality"
            )
        if isinstance(row[4].value, str) and "Four leads" in row[6].value:
            row[4].value = (
                "Go to CRM -> Leads. For each of 4 leads (next tenantN sequence), "
                "Create New Lead, fill form, Create, Convert to Deal with random Payment Type"
            )

    # --- TC-07 IMAP ---
    ws7 = wb["TC-07_ExistingContactPayment"]
    for row in ws7.iter_rows(min_row=9, max_row=ws7.max_row):
        if isinstance(row[4].value, str) and "password captured" in row[4].value:
            row[4].value = (
                "Create tenant user from Action Buttons tab; password from dialog or IMAP (Gmail app password)"
            )
        if isinstance(row[4].value, str) and "Login with generated tenant" in row[4].value:
            row[4].value = "Login with tenant credentials from IMAP or captured password (orgId from org.json)"

    # --- TC-04 IMAP if present ---
    if "TC-04_Flow1_NewOrg" in wb.sheetnames:
        ws4 = wb["TC-04_Flow1_NewOrg"]
        for row in ws4.iter_rows(min_row=9, max_row=ws4.max_row):
            if isinstance(row[4].value, str) and "YOPmail" in row[4].value:
                row[4].value = row[4].value.replace("YOPmail", "IMAP (Gmail app password)")

    # --- New sheets ---
    write_test_sheet(
        wb,
        "TC-00_Cleanup",
        "Cleanup Scenario - delete auto organizations",
        "TC-00",
        "Platform admin credentials valid: propexcel / admin@propexcel.com / Demo2026$",
        [
            ("Platform Admin", "Login at https://test.propexcel.com/login with org propexcel, admin@propexcel.com, Demo2026$", "Admin session starts"),
            ("Platform Admin", "Navigate to Organizations list (search organizations)", "Organizations page is displayed"),
            ("Platform Admin", 'Search organizations matching "Auto" and delete each (confirm dialog)', "All matching auto orgs are deleted"),
            ("Platform Admin", "Logout", "Login page is displayed"),
        ],
        "tests/Cleanup Scenario.spec.ts",
    )

    payment_steps = [
        ("Super Admin", "Login using orgId, email, password from test-data/org.json", "Admin session starts"),
        ("Super Admin", "Configure Razorpay integration and tax settings (GST 18%) if not already done", "Payment and tax setup complete"),
        ("Super Admin", "CRM -> Deals -> Create Deal with shared contact from crm-contacts-leads.json (skip lead creation)", "Deal form opens with contact selected"),
        ("Super Admin", "Add vacant property to deal, set rent/discount/GST, mark site visit done", "Deal property pricing saved"),
        ("Super Admin", "Submit for Approval and complete Deal Approve workflow", "Deal is approved"),
        ("Super Admin", "Create/view contract and approve contract", "Contract approved"),
        ("Super Admin", "Create tenant user; password from dialog or IMAP", "Tenant user created"),
        ("Super Admin", "Create move-in request and complete in Operations Requests", "Move-in completed"),
        ("Super Admin", "Logout", "Login page displayed"),
        ("Tenant", "Login with tenant credentials (IMAP or captured password)", "Tenant portal opens"),
        ("Tenant", "Logout", "Login page displayed"),
        ("Super Admin", "Create rent invoice (4000 Rental Income), submit/publish", "Invoice visible to tenant"),
        ("Super Admin", "Logout", "Login page displayed"),
        ("Tenant", "Pay invoice via Razorpay Netbanking -> Bank of Baroda -> Success", "Invoice status PAID"),
    ]

    write_test_sheet(
        wb,
        "TC-08_SkipLeadPayment",
        "Skip lead - existing contact Create Deal through rent payment (new org)",
        "TC-08",
        "TC-03 and TC-06 completed; org.json and crm-contacts-leads.json exist.",
        payment_steps,
        "tests/Skip lead flow - tenenat Payment.spec.ts",
    )

    direct_lead_steps = [
        ("Super Admin", "Login using orgId, email, password from test-data/org.json", "Admin session starts"),
        ("Super Admin", "CRM -> Leads -> Create New Lead: search tenantN name, No matches -> Create New Lead", "Lead create form opens"),
        ("Super Admin", "Fill lead form (name, email propexceltest+tenantN@gmail.com, India mobile, nationality), Create", "Lead details page opens (contact auto-created)"),
        ("Super Admin", "Convert to Deal, select payment type", "Deal Details page opens"),
        ("Super Admin", "Add property to deal, set tax, Approve deal", "Deal approved"),
        ("Super Admin", "Create/view contract and approve contract", "Contract approved"),
        ("Super Admin", "Create tenant user; password from dialog or IMAP", "Tenant user created"),
        ("Super Admin", "Create move-in request and complete in Operations", "Move-in completed"),
        ("Super Admin", "Logout", "Login page displayed"),
        ("Tenant", "Login with tenant credentials (IMAP or captured password)", "Tenant portal opens"),
        ("Tenant", "Logout", "Login page displayed"),
        ("Super Admin", "Create rent invoice (4000 Rental Income), submit/publish", "Invoice visible to tenant"),
        ("Super Admin", "Logout", "Login page displayed"),
        ("Tenant", "Pay invoice via Razorpay Netbanking -> Bank of Baroda -> Success", "Invoice status PAID"),
    ]

    write_test_sheet(
        wb,
        "TC-09_DirectLeadPayment",
        "Direct Lead - auto contact through rent payment (new org)",
        "TC-09",
        "TC-03 completed; org.json exists with valid Super Admin credentials.",
        direct_lead_steps,
        "tests/Direct Lead to Tenant Payment.spec.ts",
    )

    # Reorder sheets: Summary first, then TC-00 through TC-09
    desired_order = [
        "Summary",
        "TC-00_Cleanup",
        "TC-01_Flow1",
        "TC-02_Flow2",
        "TC-03_CreateOrganization",
        "TC-04_Flow1_NewOrg",
        "TC-05_Flow2_NewOrg",
        "TC-06_ContactsLeadsDeals",
        "TC-07_ExistingContactPayment",
        "TC-08_SkipLeadPayment",
        "TC-09_DirectLeadPayment",
    ]
    wb._sheets.sort(key=lambda s: desired_order.index(s.title) if s.title in desired_order else 99)

    wb.save(XLSX)
    print(f"Updated {XLSX}")


if __name__ == "__main__":
    main()
